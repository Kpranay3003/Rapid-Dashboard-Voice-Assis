"""
server.py  —  Rapid Dashboard Backend (Python + FastAPI)
No API key needed — chatbot is fully local in the frontend.

Install:
    pip install fastapi uvicorn openpyxl

Run:
    python server.py
"""

import openpyxl
import uvicorn
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware

# ── Load Excel once at startup ───────────────────────────────
EXCEL_PATH = "data.xlsx"
try:
    workbook = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"✅ Loaded {EXCEL_PATH}")
    print(f"   Sheets: {workbook.sheetnames}")
except FileNotFoundError:
    print(f"❌ ERROR: {EXCEL_PATH} not found. Place it next to server.py")
    workbook = None

# ── App ──────────────────────────────────────────────────────
app = FastAPI(title="Rapid Dashboard API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Helper ───────────────────────────────────────────────────
def get_sheet_data(sheet_name: str) -> list[dict]:
    if workbook is None:
        return []
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows  = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [
        str(h).strip() if h is not None else f"col_{i}"
        for i, h in enumerate(rows[0])
    ]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append({
            headers[i]: (str(v) if v is not None else "")
            for i, v in enumerate(row)
        })
    return result

# ── GET /api/node/{node_id} ──────────────────────────────────
@app.get("/api/node/{node_id}")
def get_node_data(node_id: str):
    return get_sheet_data(node_id)

# ── GET /api/summary/{node_id} ───────────────────────────────
@app.get("/api/summary/{node_id}")
def get_summary(node_id: str):
    data     = get_sheet_data(node_id)
    total    = len(data)
    success  = sum(1 for d in data if d.get("Status", "").upper() == "SUCCESS")
    failed   = sum(1 for d in data if d.get("Status", "").upper() == "FAILED")
    critical = sum(1 for d in data if d.get("CRITICAL", "").upper() == "YES")
    return {"total": total, "success": success, "failed": failed, "critical": critical}

# ── GET /api/health ──────────────────────────────────────────
@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "excel":  EXCEL_PATH if workbook else "NOT FOUND",
        "sheets": workbook.sheetnames if workbook else [],
    }

# ── Run ──────────────────────────────────────────────────────
if __name__ == "__main__":
    uvicorn.run("server:app", host="0.0.0.0", port=5000, reload=True)
